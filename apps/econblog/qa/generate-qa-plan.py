#!/usr/bin/env python3
"""Generate Adam's Axioms manual QA plan Excel workbook with status dropdown."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "ID",
    "Epic",
    "Persona",
    "User Story",
    "Priority",
    "Preconditions",
    "Test Steps",
    "Expected Result",
    "Status",
    "Tester",
    "Date Tested",
    "Notes",
]

STATUS_OPTIONS = '"Not Started,In Progress,Pass,Fail,Blocked,Skipped"'

ROWS = [
    # Epic 1: Authentication
    ("AUTH-01", "Authentication", "Anonymous", "As a visitor, I want to sign up with email/password so I can start learning.", "P0", "Logged out; on /auth?mode=signup", "1. Enter valid username, email, password (8+ chars), confirm password\n2. Submit form", "Account created; profile exists; redirected; header shows signed-in state", "Not Started", "", "", "Test duplicate username and weak password errors"),
    ("AUTH-02", "Authentication", "Anonymous", "As a visitor, I want to sign up with Google so I can onboard quickly.", "P0", "Logged out; on /auth?mode=signup", "1. Click Continue with Google\n2. Complete OAuth flow", "Account created; profile exists; redirected to intended page", "Not Started", "", "", ""),
    ("AUTH-03", "Authentication", "Free user", "As a signed-in user, I want to sign in with email/password so I can return to my progress.", "P0", "Existing account; logged out", "1. Go to /auth\n2. Enter email and password\n3. Submit", "Signed in; session active; progress available", "Not Started", "", "", ""),
    ("AUTH-04", "Authentication", "Free user", "As a signed-in user, I want to sign in with Google so I can use OAuth.", "P0", "Existing Google-linked account; logged out", "1. Go to /auth\n2. Click Continue with Google", "Signed in via OAuth; profile loads", "Not Started", "", "", ""),
    ("AUTH-05", "Authentication", "Anonymous", "As a visitor signing up, I want username validation so I pick a unique handle.", "P1", "Logged out; on signup form", "1. Enter taken username\n2. Enter invalid characters\n3. Enter valid unique username", "Errors shown for invalid/taken; valid username accepted", "Not Started", "", "", ""),
    ("AUTH-06", "Authentication", "Anonymous", "As a visitor, I want email confirmation messaging when required so I know what to do next.", "P1", "Email confirmation enabled in Supabase", "1. Sign up with new email\n2. Observe post-submit messaging", "Clear message to check email; can switch to sign-in mode", "Not Started", "", "", ""),
    ("AUTH-07", "Authentication", "Free user", "As a signed-in user, I want ?next= redirect after login so I land where I intended.", "P0", "Logged out", "1. Visit protected URL e.g. /profile?next preserved via /auth?next=/profile\n2. Sign in", "Redirected to original destination after auth", "Not Started", "", "", "Also test from /subscribe and lesson deep links"),
    ("AUTH-08", "Authentication", "Free user", "As a signed-in user, I want to sign out from the header menu so my session ends.", "P0", "Signed in", "1. Open user account menu\n2. Click Sign out", "Session cleared; header shows Sign in; protected routes redirect", "Not Started", "", "", ""),
    ("AUTH-09", "Authentication", "Free user", "As a signed-in user, I want to sign out from Profile so my session ends locally.", "P1", "Signed in; on /profile", "1. Click Sign out button on profile", "Local session cleared; redirected appropriately", "Not Started", "", "", ""),
    ("AUTH-10", "Authentication", "Anonymous", "As a visitor, I want protected routes to redirect me to auth so I cannot access gated pages.", "P0", "Logged out", "1. Visit /profile\n2. Visit /play\n3. Visit /play/price-war\n4. Visit /play/econ-wordle", "/profile and /play/* redirect to auth except Econ Wordle loads", "Not Started", "", "", ""),
    ("AUTH-11", "Authentication", "Free user", "As a signed-in user, I want my session to persist across page reloads and tabs.", "P1", "Signed in", "1. Reload page\n2. Open new tab to same site", "Session remains active in both contexts", "Not Started", "", "", ""),
    ("AUTH-12", "Authentication", "Anonymous", "As a visitor, I want a clear error when OAuth fails so I can retry.", "P2", "Logged out", "1. Trigger OAuth failure (cancel or misconfigured callback)", "Redirect to /auth?error=auth with helpful message", "Not Started", "", "", ""),
    # Epic 2: Navigation
    ("NAV-01", "Navigation", "Anonymous", "As a visitor, I want to browse the landing page so I understand the product.", "P1", "Logged out", "1. Visit /\n2. Scroll through hero, lessons, games, pricing sections", "Page loads; CTAs visible; links work", "Not Started", "", "", ""),
    ("NAV-02", "Navigation", "Anonymous", "As a visitor, I want to browse the lesson catalog with search/filters so I find content.", "P1", "Any", "1. Go to /lessons\n2. Search\n3. Filter by category/difficulty\n4. Paginate", "Results update correctly; Lesson Zero shows free; paid lessons show lock for free users", "Not Started", "", "", ""),
    ("NAV-03", "Navigation", "Anonymous", "As a visitor, I want to view the XP leaderboard so I see community progress.", "P2", "Any", "1. Go to /leaderboard", "Top users displayed; seeded + real users merged", "Not Started", "", "", ""),
    ("NAV-04", "Navigation", "Free user", "As a signed-in user, I want header nav (Lessons, Games, Profile) so I can move around.", "P0", "Signed in", "1. Click each header nav link", "Each destination loads correctly", "Not Started", "", "", ""),
    ("NAV-05", "Navigation", "Admin", "As an admin, I want an Admin link in the header so I can reach the generator.", "P1", "Signed in as admin", "1. Check header for Admin link\n2. Click it", "Admin link visible; navigates to /admin", "Not Started", "", "", ""),
    ("NAV-06", "Navigation", "Anonymous", "As a visitor, I want to view pricing on /subscribe so I can compare plans.", "P1", "Any", "1. Go to /subscribe", "Monthly and lifetime plans displayed with correct pricing", "Not Started", "", "", ""),
    # Epic 3: Lessons
    ("LES-01", "Lessons", "Anonymous", "As a visitor, I want to read Lesson Zero without an account so I can try the format.", "P0", "Logged out", "1. Go to /lessons/lesson-zero\n2. Read content", "Full lesson readable without sign-in", "Not Started", "", "", ""),
    ("LES-02", "Lessons", "Free user", "As a free user, I want full Lesson Zero access so I get a complete sample.", "P0", "Signed in as free user", "1. Open Lesson Zero\n2. Navigate all sections", "All sections accessible; quizzes available", "Not Started", "", "", ""),
    ("LES-03", "Lessons", "Free user", "As a free user, I want Section 1 preview on paid lessons so I can evaluate before subscribing.", "P0", "Signed in as free user", "1. Open a paid lesson (not Lesson Zero)", "Section 1 readable; Section 2+ locked", "Not Started", "", "", ""),
    ("LES-04", "Lessons", "Free user", "As a free user, I want a subscribe banner after Section 1 preview so I know how to unlock.", "P0", "Signed in as free user; on paid lesson Section 1", "1. Reach end of Section 1 preview", "Preview banner with CTA to /subscribe?next=...", "Not Started", "", "", ""),
    ("LES-05", "Lessons", "Member", "As a member, I want full access to all published lessons so I can complete the curriculum.", "P0", "Signed in as member", "1. Open multiple paid lessons\n2. Navigate all sections", "All sections accessible; no preview lock", "Not Started", "", "", ""),
    ("LES-06", "Lessons", "Admin", "As an admin, I want full lesson access without paying so I can QA content.", "P1", "Signed in as admin (no subscription)", "1. Open paid lesson\n2. Navigate all sections", "Full access; Generator shortcut visible on editable lessons", "Not Started", "", "", ""),
    ("LES-07", "Lessons", "Member", "As a member, I want section tabs to unlock progressively so I follow the intended sequence.", "P0", "Signed in as member; on lesson with multiple sections", "1. Complete all subsection quizzes in Section 1\n2. Check Section 2 tab", "Section 2 unlocks; incomplete sections show lock", "Not Started", "", "", ""),
    ("LES-08", "Lessons", "Free user", "As a free user, I want to be blocked from Section 2+ on paid lessons so preview is enforced.", "P0", "Signed in as free user; on paid lesson", "1. Try to access Section 2 tab or content", "Section 2+ blocked/disabled", "Not Started", "", "", ""),
    # Epic 4: Quizzes
    ("QUIZ-01", "Gated Quizzes", "Anonymous", "As a visitor on Lesson Zero, I want a sign-in prompt on quizzes so I know auth is required.", "P0", "Logged out; on Lesson Zero quiz", "1. Scroll to quiz gate", "SignInPrompt shown; cannot submit answers", "Not Started", "", "", ""),
    ("QUIZ-02", "Gated Quizzes", "Free user", "As a free user, I want to submit Lesson Zero quizzes so I can earn XP on the sample.", "P0", "Signed in as free user; on Lesson Zero", "1. Answer quiz correctly\n2. Submit", "Answer accepted; XP awarded; quiz marked complete", "Not Started", "", "", ""),
    ("QUIZ-03", "Gated Quizzes", "Free user", "As a free user, I want a subscribe prompt on paid-lesson quizzes so I cannot bypass paywall.", "P0", "Signed in as free user; on paid lesson Section 1", "1. Reach quiz on paid lesson", "SubscribePrompt shown; cannot submit", "Not Started", "", "", ""),
    ("QUIZ-04", "Gated Quizzes", "Member", "As a member, I want to answer subsection quizzes so I progress through lessons.", "P0", "Signed in as member; on paid lesson", "1. Answer quiz correctly\n2. Submit", "Answer accepted; progress advances", "Not Started", "", "", ""),
    ("QUIZ-05", "Gated Quizzes", "Member", "As a member, I want 3 attempts per question so I can learn from mistakes.", "P0", "Signed in as member; on quiz question", "1. Answer wrong (attempt 1)\n2. Answer wrong (attempt 2)\n3. Answer wrong (attempt 3)", "Penalties on attempts 1-2; 24h lockout on attempt 3", "Not Started", "", "", ""),
    ("QUIZ-06", "Gated Quizzes", "Member", "As a member, I want XP penalties on wrong answers so stakes feel real.", "P1", "Signed in as member", "1. Answer quiz wrong once\n2. Check profile XP", "XP decreases per xpPenalties config; floor at 0", "Not Started", "", "", ""),
    ("QUIZ-07", "Gated Quizzes", "Member", "As a member, I want 24h lockout after 3 wrong answers so I cannot brute-force.", "P0", "Signed in as member", "1. Fail question 3 times\n2. Try to answer again immediately", "Locked state shown; 423 on API; gold locked UI", "Not Started", "", "", ""),
    ("QUIZ-08", "Gated Quizzes", "Member", "As a member, I want to proceed past a locked question so I am not fully blocked.", "P1", "Question locked after 3 wrong", "1. Click Next after lockout", "Can advance to next subsection despite lock", "Not Started", "", "", ""),
    ("QUIZ-09", "Gated Quizzes", "Member", "As a member, I want Next disabled until quiz is correct or locked so I cannot skip checks.", "P0", "Signed in as member; unanswered quiz", "1. Try clicking Next before answering", "Next button disabled until correct or locked", "Not Started", "", "", ""),
    ("QUIZ-10", "Gated Quizzes", "Member", "As a member, I want no duplicate XP on re-submitting a correct answer.", "P1", "Question already answered correctly", "1. Re-submit same correct answer", "409 response; no additional XP", "Not Started", "", "", ""),
    ("QUIZ-11", "Gated Quizzes", "Member", "As a member, I want the mastery exam unlocked after attempting all section quizzes.", "P0", "All section quizzes attempted (not necessarily correct)", "1. Open Mastery tab", "Mastery exam available to start", "Not Started", "", "", ""),
    ("QUIZ-12", "Gated Quizzes", "Member", "As a member, I want a timed mastery exam so retention is tested under pressure.", "P1", "Mastery exam unlocked", "1. Start mastery exam\n2. Let timer run to zero", "Timer counts down; auto-submit at 0:00", "Not Started", "", "", ""),
    ("QUIZ-13", "Gated Quizzes", "Member", "As a member, I want mastery pass/fail recorded so lesson completion is accurate.", "P0", "Mastery exam completed", "1. Pass exam (70%+)\n2. Check lesson progress", "masteryPassed set; lesson marked complete; completedAt recorded", "Not Started", "", "", ""),
    ("QUIZ-14", "Gated Quizzes", "Free user", "As a free user, I want mastery exam blocked so I cannot access paid assessment.", "P0", "Signed in as free user; on paid lesson", "1. Try to access Mastery tab", "Mastery disabled/blocked", "Not Started", "", "", ""),
    # Epic 5: XP
    ("XP-01", "XP & Leaderboard", "Free user", "As a free user, I want XP on correct quiz answers so I feel rewarded.", "P0", "Signed in", "1. Answer quiz correctly\n2. Check XP display", "XP increases by xpReward; toast/update shown", "Not Started", "", "", ""),
    ("XP-02", "XP & Leaderboard", "Member", "As a member, I want XP penalties that cannot push total below 0.", "P1", "Low XP account", "1. Answer multiple questions wrong", "XP floors at 0; never negative", "Not Started", "", "", ""),
    ("XP-03", "XP & Leaderboard", "Member", "As a member, I want level = floor(totalXp/1000)+1 so progression is predictable.", "P1", "Known XP total", "1. Check profile level\n2. Earn XP to cross 1000 boundary", "Level matches formula; level-up at boundary", "Not Started", "", "", ""),
    ("XP-04", "XP & Leaderboard", "Member", "As a member, I want per-lesson XP shown in lesson sidebar so I track lesson progress.", "P1", "On lesson with quizzes", "1. Earn XP in lesson\n2. Check sidebar", "Per-lesson XP earned updates; completion rings accurate", "Not Started", "", "", ""),
    ("XP-05", "XP & Leaderboard", "Member", "As a member, I want profile Progress tab to show level bar and XP to next level.", "P1", "Signed in as member", "1. Go to /profile?tab=progress", "Level bar, XP to next level, activity shown", "Not Started", "", "", ""),
    ("XP-06", "XP & Leaderboard", "Anonymous", "As a visitor, I want to view the public leaderboard so I see top learners.", "P2", "Any", "1. Go to /leaderboard", "Top 50 displayed; usernames and XP visible", "Not Started", "", "", ""),
    ("XP-07", "XP & Leaderboard", "Member", "As a member, I want my rank on profile so I know where I stand.", "P2", "Signed in with XP", "1. Check profile Progress tab rank", "Rank displayed and reasonable vs leaderboard", "Not Started", "", "", ""),
    ("XP-08", "XP & Leaderboard", "Member", "As a member, I want level-up when crossing 1000 XP boundary mid-quiz.", "P2", "Near level boundary", "1. Earn XP pushing past 1000\n2. Observe UI", "Level increments; UI reflects new level", "Not Started", "", "", ""),
    # Epic 6: Personalization
    ("PATH-01", "Personalization", "Free user", "As a free user, I want path setup onboarding so I get a tailored sequence.", "P0", "Signed in; path not set up", "1. Trigger PathSetupModal\n2. Answer ~6 questions\n3. Complete flow", "Path saved; Profile Path tab shows ordered lessons; tagline shown", "Not Started", "", "", ""),
    ("PATH-02", "Personalization", "Free user", "As a free user, I want to skip path setup so I am not forced immediately.", "P1", "Path setup modal open", "1. Click Skip", "Modal closes; needsPathSetup remains; banner persists", "Not Started", "", "", ""),
    ("PATH-03", "Personalization", "Member", "As a member, I want my path ordered by interests so lessons feel relevant.", "P0", "Path setup completed", "1. Review Profile Path tab order", "Lessons ordered by interest weights; relevant topics first", "Not Started", "", "", ""),
    ("PATH-04", "Personalization", "Member", "As a member, I want a path tagline so personalization is visible.", "P1", "Path setup completed", "1. Check path section tagline", "Tagline reflects primary interest e.g. policy focus", "Not Started", "", "", ""),
    ("PATH-05", "Personalization", "Member", "As a member, I want path items marked completed/in_progress/up_next/locked/coming_soon.", "P0", "Path with mixed lesson states", "1. Review path timeline and list", "Each item shows correct state badge", "Not Started", "", "", ""),
    ("PATH-06", "Personalization", "Free user", "As a free user, I want non-Lesson Zero path items locked so paywall is clear.", "P0", "Signed in as free user; path set up", "1. Review path items", "Only Lesson Zero accessible; others locked", "Not Started", "", "", ""),
    ("PATH-07", "Personalization", "Member", "As a member, I want Change focus to retake onboarding so I can update interests.", "P1", "Path already set up", "1. Profile Path → Change focus\n2. New answers → save", "Path reorders; preferences overwritten; tagline updates", "Not Started", "", "", ""),
    ("PATH-08", "Personalization", "Free user", "As a free user, I want a path setup banner on /lessons until I complete or skip.", "P1", "needsPathSetup true", "1. Visit /lessons", "Banner prompts path setup", "Not Started", "", "", ""),
    ("PATH-09", "Personalization", "Member", "As a member, I want Continue on your path card so I resume easily.", "P1", "Path set up with next lesson", "1. Visit /lessons or dashboard", "Continue card links to next path lesson", "Not Started", "", "", ""),
    ("PATH-10", "Personalization", "Member", "As a member, I want unpublished lessons as coming soon on my path.", "P2", "Path includes unpublished lessons", "1. Review path list", "Unpublished items show coming soon state", "Not Started", "", "", ""),
    # Epic 7: Profile
    ("PROF-01", "Profile", "Free user", "As a signed-in user, I want to view my profile so I see my identity and progress.", "P0", "Signed in", "1. Go to /profile", "Profile loads with username, avatar, tabs", "Not Started", "", "", ""),
    ("PROF-02", "Profile", "Free user", "As a signed-in user, I want to change my username so I can personalize my account.", "P1", "Signed in", "1. Profile Personal tab → change username → save", "Username updates; uniqueness validated", "Not Started", "", "", ""),
    ("PROF-03", "Profile", "Free user", "As a signed-in user, I want an avatar (generated or Google) so my profile feels personal.", "P2", "Signed in", "1. Check avatar on profile and header", "Avatar displays correctly for auth method", "Not Started", "", "", ""),
    ("PROF-04", "Profile", "Member", "As a member, I want Profile tabs via ?tab= deep links.", "P1", "Signed in as member", "1. Visit /profile?tab=path\n2. ?tab=subscription\n3. ?tab=progress", "Correct tab opens for each deep link", "Not Started", "", "", ""),
    ("PROF-05", "Profile", "Member", "As a member, I want Progress tab with completed lessons, heatmap, streak.", "P1", "Member with lesson activity", "1. Open Progress tab", "Completed lessons, heatmap, streak days shown", "Not Started", "", "", ""),
    ("PROF-06", "Profile", "Member", "As a member, I want Subscription tab showing plan status.", "P0", "Active subscription", "1. Open Subscription tab", "Plan type, status, manage billing CTA shown", "Not Started", "", "", ""),
    ("PROF-07", "Profile", "Lapsed member", "As a lapsed subscriber, I want subscription status to reflect canceled/past_due.", "P0", "Canceled or past_due subscription", "1. Open Subscription tab\n2. Try paid lesson access", "Status shows lapsed; lesson access revoked", "Not Started", "", "", ""),
    # Epic 8: Billing
    ("BILL-01", "Billing", "Anonymous", "As a visitor, I want to view plans on /subscribe.", "P1", "Any", "1. Go to /subscribe", "Monthly and lifetime plans with pricing displayed", "Not Started", "", "", ""),
    ("BILL-02", "Billing", "Free user", "As a free user, I want checkout to require auth so billing is tied to my account.", "P0", "Logged out", "1. Go to /subscribe\n2. Click subscribe on a plan", "Redirected to /auth?next=...", "Not Started", "", "", ""),
    ("BILL-03", "Billing", "Free user", "As a free user, I want Stripe checkout for monthly/lifetime so I can subscribe.", "P0", "Signed in as free user; Stripe test mode", "1. Select plan\n2. Complete Stripe test checkout", "Subscription active; member access unlocked", "Not Started", "", "", ""),
    ("BILL-04", "Billing", "Free user", "As a free user, I want ?next= preserved through checkout so I return to intended lesson.", "P0", "On paid lesson preview", "1. Click subscribe CTA with next param\n2. Complete checkout", "Returned to original lesson with full access", "Not Started", "", "", ""),
    ("BILL-05", "Billing", "Member", "As a member, I want immediate lesson access after successful payment.", "P0", "Just completed checkout", "1. Return from Stripe\n2. Open previously locked lesson", "Full lesson access without manual refresh delay", "Not Started", "", "", "Note webhook timing edge case"),
    ("BILL-06", "Billing", "Member", "As a member, I want Stripe billing portal to manage/cancel subscription.", "P1", "Active subscription", "1. Profile Subscription → Manage billing\n2. Open portal", "Stripe portal loads; can view/cancel plan", "Not Started", "", "", ""),
    ("BILL-07", "Billing", "Free user", "As a free user, I want canceled checkout to return gracefully.", "P2", "In Stripe checkout", "1. Cancel/back out of checkout", "Return to site with ?canceled=1; no broken state", "Not Started", "", "", ""),
    ("BILL-08", "Billing", "Lapsed member", "As a lapsed subscriber, I want lesson access revoked after cancellation.", "P0", "Subscription canceled", "1. Open paid lesson\n2. Try quiz submit", "Preview mode restored; SubscribePrompt on quizzes", "Not Started", "", "", ""),
    # Epic 9: Wordle
    ("WORD-01", "Econ Wordle", "Anonymous", "As a visitor, I want to play today's Wordle without an account.", "P0", "Logged out", "1. Go to /play/econ-wordle", "Game loads; no auth redirect", "Not Started", "", "", ""),
    ("WORD-02", "Econ Wordle", "Anonymous", "As a visitor, I want 6 guesses and on-screen keyboard.", "P0", "On Wordle page", "1. Type guesses via keyboard\n2. Submit up to 6 times", "Guesses accepted; keyboard highlights correct/present/absent", "Not Started", "", "", ""),
    ("WORD-03", "Econ Wordle", "Anonymous", "As a visitor, I want local streak/progress persisted in browser.", "P1", "Played at least one game", "1. Reload page\n2. Check streak/progress", "Local storage retains progress and streak", "Not Started", "", "", ""),
    ("WORD-04", "Econ Wordle", "Anonymous", "As a visitor, I want hints after N guesses.", "P2", "Multiple wrong guesses", "1. Continue guessing until hint threshold", "Hint unlocks; word length may reveal", "Not Started", "", "", ""),
    ("WORD-05", "Econ Wordle", "Anonymous", "As a visitor, I want term + definition on finish.", "P1", "Game won or lost", "1. Complete puzzle", "Term and definition displayed on finish screen", "Not Started", "", "", ""),
    ("WORD-06", "Econ Wordle", "Anonymous", "As a visitor, I want Learn this term lesson link.", "P1", "Game finished", "1. Click lesson link on finish screen", "Navigates to related lesson", "Not Started", "", "", ""),
    ("WORD-07", "Econ Wordle", "Anonymous", "As a visitor, I want emoji share grid.", "P2", "Game finished", "1. Click share\n2. Copy grid", "Emoji result grid copied to clipboard", "Not Started", "", "", ""),
    ("WORD-08", "Econ Wordle", "Admin", "As an admin, I want debug affordances on Wordle if applicable.", "P3", "Signed in as admin", "1. Play Wordle as admin", "Admin-only debug UI visible if implemented", "Not Started", "", "", ""),
    # Epic 10: Margin core
    ("MGN-01", "Margin", "Anonymous", "As a visitor, I want /play to require auth so Margin is account-gated.", "P0", "Logged out", "1. Visit /play", "Redirect to auth", "Not Started", "", "", ""),
    ("MGN-02", "Margin", "Free user", "As a free user, I want the Games hub listing Wordle + Margin.", "P0", "Signed in", "1. Go to /play", "Games catalog shows Econ Wordle and Margin cards", "Not Started", "", "", ""),
    ("MGN-03", "Margin", "Free user", "As a free user, I want the tutorial vs Prof. Aldo so I learn rules.", "P0", "Signed in as free user", "1. Go to /play/price-war/tutorial\n2. Complete tutorial flow", "Tutorial match runs; coach guidance shown; can submit moves", "Not Started", "", "", ""),
    ("MGN-04", "Margin", "Free user", "As a free user, I want vs-bot matches so I can practice solo.", "P0", "Signed in as free user", "1. Lobby → Play vs bot\n2. Play through match", "Bot match created; full flow works to post-match", "Not Started", "", "", ""),
    ("MGN-05", "Margin", "Free user", "As a free user, I want Rapid 15-min mode so I can play timed matches.", "P0", "Signed in as free user", "1. Select Rapid 15+0\n2. Queue or play", "Rapid mode available and playable", "Not Started", "", "", "Verify against product spec: Rapid free, Blitz paid"),
    ("MGN-06", "Margin", "Free user", "As a free user, I want Blitz 5-min blocked so fast mode is member-only.", "P0", "Signed in as free user", "1. Try to select/play Blitz 5+0", "Blitz blocked with 403 or UI lock", "Not Started", "", "", ""),
    ("MGN-07", "Margin", "Member", "As a member, I want Blitz 5-min mode so I can play fast games.", "P0", "Signed in as member", "1. Select Blitz 5+0\n2. Queue and play", "Blitz available and playable", "Not Started", "", "", ""),
    ("MGN-08", "Margin", "Free user", "As a free user, I want only 1 concurrent match so caps are enforced.", "P0", "Free user with match in progress", "1. Start second match while first active", "Second match blocked with cap error message", "Not Started", "", "", ""),
    ("MGN-09", "Margin", "Member", "As a member, I want up to 5 concurrent matches.", "P1", "Signed in as member", "1. Start multiple matches (up to 5)", "Up to 5 concurrent; 6th blocked", "Not Started", "", "", ""),
    ("MGN-10", "Margin", "Free/Member", "As a player, I want full match flow: lobby → briefing → decide → review → waiting → report → continue.", "P0", "Signed in", "1. Start match\n2. Play full round cycle through post-match", "All phases work; URLs sync with server phase", "Not Started", "", "", ""),
    ("MGN-11", "Margin", "Free/Member", "As a player, I want legal moves enforced so invalid actions are rejected.", "P1", "In decide phase", "1. Attempt invalid move combinations", "Only legal moves selectable; API rejects illegal submits", "Not Started", "", "", ""),
    ("MGN-12", "Margin", "Free/Member", "As a player, I want round reports showing economic outcomes.", "P1", "Match in progress", "1. Complete a round\n2. View report", "Economic metrics and outcomes displayed per round", "Not Started", "", "", ""),
    ("MGN-13", "Margin", "Free/Member", "As a player, I want post-match summary with outcome.", "P0", "Match completed", "1. Reach post-match screen", "Winner/loser, key stats, play again CTA shown", "Not Started", "", "", ""),
    ("MGN-14", "Margin", "Member", "As a member, I want rated/Elo matches when MARGIN_RATED_ENABLED is on.", "P1", "Rated flag enabled; member account", "1. Play ranked queue match\n2. Check rating delta", "Elo/rating updates on post-match", "Not Started", "", "", ""),
    ("MGN-15", "Margin", "Free user", "As a free user, I want unrated messaging when rated is enabled.", "P2", "Rated flag on; free account", "1. Play Rapid match as free user", "Unrated label/message shown appropriately", "Not Started", "", "", ""),
    # Epic 11: Margin edge cases
    ("MGN-16", "Margin", "Free/Member", "As a player, I want timeout forfeit when clock runs out.", "P1", "Match with low clock", "1. Let clock expire without submitting", "Forfeit/timeout outcome; match ends", "Not Started", "", "", ""),
    ("MGN-17", "Margin", "Free/Member", "As a player, I want inactivity forfeit after idle rounds.", "P2", "Match in progress", "1. Skip activity for configured idle rounds", "Inactivity forfeit triggered", "Not Started", "", "", ""),
    ("MGN-18", "Margin", "Free/Member", "As a player, I want bankruptcy terminal screen when applicable.", "P1", "Match leading to bankruptcy", "1. Play until bankruptcy condition", "Bankruptcy screen shown; match terminal", "Not Started", "", "", ""),
    ("MGN-19", "Margin", "Free/Member", "As a player, I want abandoned match handling when opponent leaves.", "P1", "PvP match", "1. Opponent abandons or disconnects", "Abandoned outcome handled gracefully", "Not Started", "", "", ""),
    ("MGN-20", "Margin", "Free/Member", "As a player, I want mid-match refresh to recover to correct phase.", "P1", "Match in progress", "1. Refresh browser mid-match", "Loading gate then correct phase URL restored", "Not Started", "", "", ""),
    ("MGN-21", "Margin", "Free user", "As a free user, I want template-only post-match coach debrief.", "P1", "Free user; match completed", "1. View post-match coach section", "Template debrief shown; no LLM coach", "Not Started", "", "", ""),
    ("MGN-22", "Margin", "Member", "As a member, I want AI coach debrief when LLM is available.", "P1", "Member; OpenRouter configured", "1. Complete match\n2. View coach debrief", "LLM-generated personalized debrief shown", "Not Started", "", "", ""),
    ("MGN-23", "Margin", "Free/Member", "As a player, I want recommended lessons after match.", "P1", "Match completed", "1. Check post-match recommended lessons", "Lesson links shown; Start lesson navigates correctly", "Not Started", "", "", ""),
    ("MGN-24", "Margin", "Free/Member", "As a player, I want match history listing past games.", "P1", "At least one completed match", "1. Go to /play/price-war/history", "Past matches listed with outcomes and dates", "Not Started", "", "", ""),
    ("MGN-25", "Margin", "Member", "As a member, I want Margin leaderboard when rated is on.", "P2", "Rated enabled", "1. Go to /play/price-war/leaderboard", "Rankings displayed", "Not Started", "", "", ""),
    ("MGN-26", "Margin", "Free/Member", "As a player, I want notifications page for match events.", "P2", "Match events occurred", "1. Go to /play/price-war/notifications", "Notifications listed", "Not Started", "", "", ""),
    # Epic 12: Admin
    ("ADM-01", "Admin", "Admin", "As an admin, I want access to /admin lesson generator.", "P1", "Signed in as admin", "1. Go to /admin", "Admin dashboard loads with lesson generator", "Not Started", "", "", ""),
    ("ADM-02", "Admin", "Non-admin", "As a non-admin, I want admin routes blocked.", "P1", "Signed in as free/member (not admin)", "1. Visit /admin", "Redirect with admin_forbidden error", "Not Started", "", "", ""),
    ("ADM-03", "Admin", "Admin", "As an admin, I want the lesson pipeline Sources through Publish.", "P2", "Admin access; test lesson", "1. Walk pipeline tabs\n2. Generate/preview content", "Each tab functional; preview uses LessonPlayer", "Not Started", "", "", ""),
    ("ADM-04", "Admin", "Admin", "As an admin, I want Margin admin: matches, players, costs, move catalog.", "P2", "Signed in as admin", "1. Visit /admin/pricewar\n2. Check each section", "Match traces, player search, costs, move catalog load", "Not Started", "", "", ""),
    ("ADM-05", "Admin", "Admin", "As an admin, I want Generator shortcut on editable lessons.", "P3", "Admin on lesson page", "1. Open lesson as admin", "Generator link visible and navigates to editor", "Not Started", "", "", ""),
    # Epic 13: Cross-cutting
    ("X-01", "Cross-cutting", "All", "As a user, I want mobile-responsive lesson tabs and quiz UI.", "P1", "Mobile viewport or device", "1. Open lesson on mobile\n2. Navigate tabs and quizzes", "Tabs scroll; quiz stacks; usable on small screens", "Not Started", "", "", ""),
    ("X-02", "Cross-cutting", "All", "As a user, I want mobile-responsive Margin screens.", "P1", "Mobile viewport", "1. Play Margin match on mobile", "All phases usable; loading gates work", "Not Started", "", "", ""),
    ("X-03", "Cross-cutting", "All", "As a user, I want Margin disabled gracefully when PRICEWAR_ENABLED=false.", "P2", "Feature flag off", "1. Try to access /play/price-war", "Redirect to /lessons?notice=pricewar_disabled or 503", "Not Started", "", "", ""),
    ("X-04", "Cross-cutting", "Member", "As a member, I want lesson SEO/metadata on public pages.", "P3", "Any", "1. View page source on lesson page", "Title, meta, JSON-LD present", "Not Started", "", "", ""),
    ("X-05", "Cross-cutting", "Free user", "As a free user, I want Wordle to lesson bridge working end-to-end.", "P1", "Wordle completed", "1. Click Learn this term link\n2. Complete lesson flow", "Navigation works; lesson loads correctly", "Not Started", "", "", ""),
    ("X-06", "Cross-cutting", "Member", "As a member, I want game to recommended lesson bridge end-to-end.", "P1", "Match completed", "1. Click recommended lesson from post-match\n2. Start lesson", "Lesson opens; progress tracked", "Not Started", "", "", ""),
]

OUTPUT = "ADAMS_AXIOMS_QA_PLAN.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "QA Plan"

    # Header row
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Data rows
    for row_idx, row_data in enumerate(ROWS, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP
            cell.border = BORDER
            if col_idx == 5:  # Priority
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if col_idx == 9:  # Status
                cell.alignment = Alignment(horizontal="center", vertical="top")

    # Status dropdown (column I = 9)
    status_col = get_column_letter(9)
    dv = DataValidation(
        type="list",
        formula1=STATUS_OPTIONS,
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid status",
        error="Choose: Not Started, In Progress, Pass, Fail, Blocked, or Skipped",
    )
    dv.add(f"{status_col}2:{status_col}{len(ROWS) + 1}")
    ws.add_data_validation(dv)

    # Column widths
    widths = [10, 18, 14, 42, 8, 28, 36, 32, 14, 14, 14, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

    # Reference sheet: personas & status legend
    ref = wb.create_sheet("Reference")
    ref["A1"] = "Test Personas"
    ref["A1"].font = Font(bold=True, size=12)
    personas = [
        ("Anonymous", "No account — public browsing, Wordle, Lesson Zero read"),
        ("Free user", "Signed up, no subscription"),
        ("Member", "Active monthly or lifetime subscription"),
        ("Member (monthly)", "Active monthly Stripe subscription"),
        ("Member (lifetime)", "Active lifetime subscription"),
        ("Lapsed member", "Canceled or past_due subscription"),
        ("Admin", "Email in ADMIN_EMAILS env var"),
        ("Non-admin", "Signed-in user without admin email"),
        ("Free/Member", "Applies to both free and paid players"),
        ("All", "Any persona including anonymous"),
    ]
    for i, (name, desc) in enumerate(personas, 3):
        ref.cell(row=i, column=1, value=name).font = Font(bold=True)
        ref.cell(row=i, column=2, value=desc)

    ref["A14"] = "Status Values"
    ref["A14"].font = Font(bold=True, size=12)
    statuses = [
        ("Not Started", "Test case not yet run"),
        ("In Progress", "Currently being tested"),
        ("Pass", "All expected results met"),
        ("Fail", "Bug found or expected result not met"),
        ("Blocked", "Cannot test — dependency or environment issue"),
        ("Skipped", "Intentionally not tested this cycle"),
    ]
    for i, (status, desc) in enumerate(statuses, 16):
        ref.cell(row=i, column=1, value=status).font = Font(bold=True)
        ref.cell(row=i, column=2, value=desc)

    ref["A23"] = "Priority"
    ref["A23"].font = Font(bold=True, size=12)
    for i, (p, desc) in enumerate([
        ("P0", "Smoke / release blocker — must pass"),
        ("P1", "Important — should pass before release"),
        ("P2", "Nice to have — test if time allows"),
        ("P3", "Low priority / admin-only"),
    ], 25):
        ref.cell(row=i, column=1, value=p).font = Font(bold=True)
        ref.cell(row=i, column=2, value=desc)

    ref.column_dimensions["A"].width = 22
    ref.column_dimensions["B"].width = 55

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} ({len(ROWS)} test cases)")


if __name__ == "__main__":
    main()
